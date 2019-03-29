
# !/usr/bin/python
# -*- coding: utf-8 -*-
"""
-------------------------------------------------
   File Name：     do_excel.py  
   Description :  
   Author :        luoyunan
   date：          2019/03/27
   Copyright:      (c) luoyunan 2019
-------------------------------------------------
   Change Activity:
                   2019/03/27: 
-------------------------------------------------
"""

from openpyxl import load_workbook


class DoExcel:
    def __init__(self, file_name, sheet_name):
        self.file_name= file_name
        self.sheet_name= sheet_name

    def read_data(self):
        wb=load_workbook(self.file_name)
        sheet=load_workbook(self.sheet_name)

        test_data=[]
        for i in range(2, sheet.max_row+1):

            row_data={}

            row_data["CaseId"]=sheet.cell(i,1).value
            row_data["Module"]=sheet.cell(i,2).value
            row_data["Title "]=sheet.cell(i,3).value
            row_data["Url"]=sheet.cell(i,4).value
            row_data["Method"]=sheet.cell(i,5).value
            row_data["Data"] = sheet.cell(i, 6).value
            row_data["Headers"] = sheet.cell(i, 7).value
            row_data["ExpectedResult"] = sheet.cell(i, 8).value
            row_data["ActulResult"] = sheet.cell(i, 9).value

            test_data.append(row_data)

        wb.close()
        return test_data


    def write_back(self,row,col,value):
        wb=load_workbook(self.file_name)
        sheet=load_workbook(self.sheet_name)

        sheet.cell(row, col).value = value
        wb.save(self.file_name)
        wb.close()





if __name__ == '__main__':
    file_name= "test_api.xlsx"
    sheet_name= "test_case"
    test_date= DoExcel(file_name,sheet_name).read_data()
    print(test_date)

